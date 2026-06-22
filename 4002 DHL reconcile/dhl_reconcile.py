#!/usr/bin/env python3
"""
DHL invoice reconciliation.

Usage:
    python3 dhl_reconcile.py "/path/to/folder/of/pdfs"

Outputs (same folder as input):
    dhl_reconcile_lines.csv   one row per invoice, matches the reconcile xlsx layout
    dhl_bank_statement.csv    totals per GL category, matches DHLBankStatementImport.csv
    dhl_review.csv            only written if any invoice failed a sanity check
"""

import sys
import re
import csv
import shutil
import pathlib
from collections import defaultdict
from datetime import date, datetime

import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


# ---------- Configuration ----------

EU_COUNTRIES = {
    "AT","BE","BG","HR","CY","CZ","DK","EE","FI","FR","DE","GR","HU","IE","IT",
    "LV","LT","LU","MT","NL","PL","PT","RO","SK","SI","ES","SE",
}

# Reconcile CSV columns in the order they appear in the xlsx
RECONCILE_COLUMNS = [
    "Amount",
    "Check",
    "Freight Amounts",                  # 7070
    "Customer Delivery Fees",           # 7060
    "Duties & Brokerage Drawbacks",     # 7025
    "Duties & Brokerage",               # 7020
    "GST",                              # Tax Adjustment
    "EU VAT",                           # 2250
    "UK VAT",                           # 2240
    "Invoice Number",
    "Origin",
    "Notes",
]


def money(s):
    if s is None or str(s).strip() == "":
        return 0.0
    return float(str(s).replace(",", "").strip())


def round2(x):
    return round(float(x), 2)


# ---------- Format detection ----------

def detect_format(text):
    if "CUSTOMS DUTY INVOICE" in text:
        return "E10"
    if "INBOUND INVOICE" in text:
        # The INBOUND template wraps two different Types of Service. Route by it:
        # DUTIES & TAXES → duty/VAT/brokerage; EXPRESS WORLDWIDE NONDOC → delivery freight.
        if "DUTIES & TAXES" in text:
            return "YVRIR"
        if "EXPRESS WORLDWIDE" in text:
            return "YVRR_INBOUND"
        return "YVRIR"
    if "OUTBOUND INVOICE" in text:
        return "YVRR"
    return None


# ---------- Parsers ----------

def parse_meta(full_text):
    """Common header extraction."""
    inv_no = re.search(r"Invoice Number:\s*(\S+)", full_text)
    inv_dt = re.search(r"Invoice Date:\s*(\S+)", full_text)
    total = re.search(r"Total Amount \(CAD\)?:?\s*([\d,.]+)", full_text)
    return {
        "invoice_number": inv_no.group(1) if inv_no else "",
        "invoice_date":   inv_dt.group(1) if inv_dt else "",
        "total":          money(total.group(1)) if total else 0.0,
    }


def parse_yvrir(full_text):
    """INBOUND INVOICE: outbound shipments billed as DUTIES & TAXES.
    Returns dict with per-shipment charge breakdown so we can route by country."""
    inv = parse_meta(full_text)
    inv["format"] = "YVRIR"

    # Split into shipment blocks. Each starts with a line: <AWB> <ref> <date> ...
    blocks = re.split(r"\n(?=\d{10}\s+\S+\s+\d{2}/\d{2}/\d{4})", full_text)
    shipments = []
    for block in blocks:
        if not re.match(r"^\d{10}\s+\S+\s+\d{2}/\d{2}/\d{4}", block):
            continue
        awb = re.match(r"^(\d{10})", block).group(1)

        # Destination country: find all XX-... patterns, take the first non-CA.
        # Match "DE-21031", "GB-PO16", "MX-07140", etc.
        country_matches = re.findall(r"\b([A-Z]{2})-[A-Z0-9]", block)
        dest_country = next((c for c in country_matches if c != "CA"), None)

        # Charges — known DHL line items used on YVRIR invoices.
        charges = {}
        charge_names = [
            "IMPORT EXPORT TAXES",
            "IMPORT EXPORT DUTIES",
            "DUTY TAX PAID",
            "CLEARANCE PROCESSING",
            "EXCISE TAX",
            "OTHER GOVT DEPT FEE",
            "OTHER GOVERNMENT DEPT FEE",
            "NON-ROUTINE ENTRY",
            "BONDED STORAGE",
            "REGULATORY CHARGES",
        ]
        for name in charge_names:
            m = re.search(rf"{re.escape(name)}\s+([\d.,]+)\s+([\d.,]+)", block)
            if m:
                # the duplicated amount is excl-tax then incl-tax; for YVRIR they match
                charges[name] = money(m.group(2))

        shipments.append({"awb": awb, "country": dest_country, "charges": charges})

    inv["shipments"] = shipments
    return inv


def parse_yvrr(full_text):
    """OUTBOUND INVOICE: EXPRESS WORLDWIDE NONDOC — everything → Customer Delivery Fees.
    We pull DUTIES TAXES PAID separately in case it ever appears."""
    inv = parse_meta(full_text)
    inv["format"] = "YVRR"

    # Optional: detect Duties Taxes Paid line in the summary if it ever appears
    m = re.search(r"DUTIES?\s+TAX(?:ES)?\s+PAID\s+([\d.,]+)", full_text)
    inv["duties_taxes_paid"] = money(m.group(1)) if m else 0.0
    return inv


def parse_yvrr_inbound(full_text):
    """INBOUND template but Type of Service is EXPRESS WORLDWIDE NONDOC.
    Freight charge → Customer Delivery Fees; GST (if any) split out separately."""
    inv = parse_meta(full_text)
    inv["format"] = "YVRR_INBOUND"
    # When GST applies the totals line reads "Total Amount (CAD) <excl> <tax> <incl>".
    m = re.search(r"Total Amount \(CAD\)\s+([\d.,]+)(?:\s+([\d.,]+)\s+([\d.,]+))?", full_text)
    if m and m.group(3):
        inv["excl_tax"] = money(m.group(1))
        inv["gst"]      = money(m.group(2))
        inv["total"]    = money(m.group(3))
    else:
        inv["excl_tax"] = inv["total"]
        inv["gst"]      = 0.0
    return inv


def parse_e10(full_text):
    """CUSTOMS DUTY INVOICE: import to Canada.
    Charges split into two sections (Duties/Taxes/Regulatory and DHL Charges)."""
    inv = parse_meta(full_text)
    inv["format"] = "E10"
    # E10 has different total label
    if not inv["total"]:
        m = re.search(r"Total Amount \(CAD\):\s*([\d.,]+)", full_text)
        if m:
            inv["total"] = money(m.group(1))

    # Origin code (3-letter): line like "...HKG YVR 1 2.63"
    m = re.search(r"\b([A-Z]{3})\s+YVR\s+\d", full_text)
    inv["origin_code"] = m.group(1) if m else ""

    # AWB (waybill 10-digit) — first standalone 10-digit before a description
    m = re.search(r"\n(\d{10})\b", full_text)
    inv["awb"] = m.group(1) if m else ""

    # Walk lines, track section
    line_re = re.compile(
        r"^(.+?)\s+(\d+\.\d{2})\s+([A-Za-z/]+)\s+(\d+\.\d{2})\s+(\d+\.\d{2})\s*$"
    )
    charges = {"regulatory": [], "dhl": []}
    section = None
    for line in full_text.splitlines():
        s = line.strip()
        if "Duties, Taxes and Regulatory Charges" in s:
            section = "regulatory"; continue
        if s == "DHL Charges" or s.startswith("DHL Charges"):
            section = "dhl"; continue
        if s.startswith("Total ") or s.startswith("Analysis of TAX"):
            section = None; continue
        if not section:
            continue
        m = line_re.match(s)
        if m:
            charges[section].append({
                "name": m.group(1).strip(),
                "excl_tax": money(m.group(2)),
                "tax": money(m.group(4)),
                "incl_tax": money(m.group(5)),
            })
    inv["charges"] = charges
    return inv


# ---------- Allocation rules ----------

def allocate(inv):
    """Return (allocation_dict, notes_list). Allocation values sum to invoice total."""
    alloc = defaultdict(float)
    notes = []

    if inv["format"] == "YVRIR":
        # Per-shipment, country-aware routing
        for s in inv["shipments"]:
            country = s["country"]
            for name, amt in s["charges"].items():
                if name == "IMPORT EXPORT TAXES":
                    if country in EU_COUNTRIES:
                        alloc["EU VAT"] += amt
                    elif country == "GB":
                        alloc["UK VAT"] += amt
                    else:
                        alloc["Duties & Brokerage"] += amt
                        if not country:
                            notes.append(f"AWB {s['awb']}: no country detected; IET → D&B")
                else:
                    # Duties, Duty Tax Paid, Clearance Processing, Excise, OGD fees
                    alloc["Duties & Brokerage"] += amt

    elif inv["format"] == "YVRR":
        # Everything → Customer Delivery Fees (Fuel, Standard, GoGreen, Remote, Emergency...)
        # Exception: if Duties Taxes Paid appears, that goes to Duties & Brokerage
        dtp = inv.get("duties_taxes_paid", 0.0)
        alloc["Customer Delivery Fees"] += inv["total"] - dtp
        if dtp > 0:
            alloc["Duties & Brokerage"] += dtp

    elif inv["format"] == "YVRR_INBOUND":
        # Express freight on the inbound template → Customer Delivery Fees, GST split out.
        alloc["Customer Delivery Fees"] += inv["excl_tax"]
        if inv.get("gst"):
            alloc["GST"] += inv["gst"]

    elif inv["format"] == "E10":
        # Import to Canada (CUSTOMS DUTY INVOICE → ST MORITZ WATCH CORP, YVR).
        # All tax amounts → GST. All excl-tax amounts (incl duties, clearance fees) → Freight Amounts.
        for c in inv["charges"]["regulatory"] + inv["charges"]["dhl"]:
            alloc["GST"] += c["tax"]
            alloc["Freight Amounts"] += c["excl_tax"]

    # Sanity check
    allocated = round2(sum(alloc.values()))
    expected = round2(inv["total"])
    if abs(allocated - expected) > 0.02:
        notes.append(f"Allocation ${allocated:.2f} != invoice total ${expected:.2f}")

    return alloc, notes


# ---------- Main ----------

def main(folder):
    folder = pathlib.Path(folder).expanduser()
    if not folder.is_dir():
        print(f"Not a folder: {folder}", file=sys.stderr); sys.exit(2)

    pdfs = sorted(folder.glob("*.pdf"))
    if not pdfs:
        print(f"No PDFs found in {folder}", file=sys.stderr); sys.exit(2)

    rows = []
    category_totals = defaultdict(float)
    reviews = []

    for pdf_path in pdfs:
        try:
            with pdfplumber.open(pdf_path) as pdf:
                full_text = "\n".join(p.extract_text() or "" for p in pdf.pages)
            fmt = detect_format(full_text)
            if fmt == "YVRIR":  inv = parse_yvrir(full_text)
            elif fmt == "YVRR": inv = parse_yvrr(full_text)
            elif fmt == "YVRR_INBOUND": inv = parse_yvrr_inbound(full_text)
            elif fmt == "E10":  inv = parse_e10(full_text)
            else:
                reviews.append({"file": pdf_path.name, "issue": "unrecognized invoice format"})
                continue
        except Exception as e:
            reviews.append({"file": pdf_path.name, "issue": f"parse error: {e}"})
            continue

        alloc, notes = allocate(inv)

        row = {col: "" for col in RECONCILE_COLUMNS}
        row["Amount"] = round2(inv["total"])
        row["Check"] = 0
        for k in ["Freight Amounts", "Customer Delivery Fees",
                  "Duties & Brokerage Drawbacks", "Duties & Brokerage",
                  "GST", "EU VAT", "UK VAT"]:
            v = alloc.get(k, 0)
            row[k] = round2(v) if v else ""
        row["Invoice Number"] = inv["invoice_number"]
        row["Origin"] = inv.get("origin_code", "") if inv["format"] == "E10" else ""
        row["Notes"] = "; ".join(notes)
        rows.append(row)

        for k, v in alloc.items():
            category_totals[k] += v
        if notes:
            reviews.append({"file": pdf_path.name, "issue": "; ".join(notes)})

    # ===== Reconcile XLSX (matches DHL Reconcile.xlsx structure) =====
    out_xlsx = folder / "DHL Reconcile.xlsx"

    # Back up existing file if present
    if out_xlsx.exists():
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = out_xlsx.with_suffix(f".bak-{ts}.xlsx")
        shutil.copy2(out_xlsx, backup)
        print(f"Backed up existing xlsx → {backup.name}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    bold = Font(bold=True)
    yellow_fill = PatternFill("solid", fgColor="FFFF00")

    # Row 1: column headers
    headers = ["Amount", "Check", "Freight Amounts", "Customer Delivery Fees",
               "Duties & Brokerage (Drawbacks)", "Duties & Brokerage",
               "GST", "EU VAT", "UK VAT"]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = bold
        c.fill = yellow_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # Row 2: GL codes (matches existing xlsx)
    gl_codes = ["", "", "7070", "7060", "7025", "7020", "Tax Adjustment", "2250", "2240"]
    for col_idx, code in enumerate(gl_codes, 1):
        ws.cell(row=2, column=col_idx, value=code)

    # Rows 3+: per-invoice rows
    col_map = {
        "Amount": 1, "Check": 2, "Freight Amounts": 3, "Customer Delivery Fees": 4,
        "Duties & Brokerage Drawbacks": 5, "Duties & Brokerage": 6,
        "GST": 7, "EU VAT": 8, "UK VAT": 9,
    }
    r_idx = 3
    for row in rows:
        for col_name, c_idx in col_map.items():
            if col_name == "Check":
                continue  # Check is a formula, set below
            v = row.get(col_name, "")
            if v == "" or v is None or v == 0:
                continue
            ws.cell(row=r_idx, column=c_idx, value=v)
        # Column B: live check formula. 0 = balanced; non-zero = look at this row.
        ws.cell(row=r_idx, column=2, value=f"=A{r_idx}-SUM(C{r_idx}:I{r_idx})")
        # Columns J (AWB) and K (Origin) for inbound rows
        if row["Origin"]:
            ws.cell(row=r_idx, column=11, value=row["Origin"])
        ws.cell(row=r_idx, column=12, value=row["Invoice Number"])
        if row["Notes"]:
            ws.cell(row=r_idx, column=13, value=row["Notes"])
        r_idx += 1

    # Totals row
    total_row = r_idx
    first_data_row = 3
    last_data_row = total_row - 1
    ws.cell(row=total_row, column=1, value="Total").font = bold
    # B (totals row) = grand total of column A (matches your previous reconcile pattern)
    ws.cell(row=total_row, column=2,
            value=f"=SUM(A{first_data_row}:A{last_data_row})").font = bold
    for col_name, c_idx in col_map.items():
        if col_name in ("Amount", "Check"):
            continue
        tot = sum((r.get(col_name) or 0) for r in rows)
        if tot:
            cell = ws.cell(row=total_row, column=c_idx, value=round2(tot))
            cell.font = bold

    # Auto-width columns
    for col_letter, width in zip("ABCDEFGHIJKLM",
                                  [10, 10, 14, 18, 16, 14, 10, 10, 10, 14, 8, 18, 30]):
        ws.column_dimensions[col_letter].width = width

    # Header labels for J/K/L/M (extra columns)
    ws.cell(row=1, column=10, value="").font = bold
    ws.cell(row=1, column=11, value="Origin").font = bold
    ws.cell(row=1, column=12, value="Invoice #").font = bold
    ws.cell(row=1, column=13, value="Notes").font = bold

    wb.save(out_xlsx)
    print(f"Wrote: {out_xlsx.name}")

    # Also keep CSV for paste convenience
    out_reconcile = folder / "dhl_reconcile_lines.csv"
    with open(out_reconcile, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=RECONCILE_COLUMNS)
        w.writeheader()
        for row in rows:
            w.writerow(row)
        totals = {c: "" for c in RECONCILE_COLUMNS}
        totals["Amount"] = round2(sum(r["Amount"] for r in rows))
        totals["Check"] = ""
        totals["Invoice Number"] = "TOTAL"
        for col in ["Freight Amounts", "Customer Delivery Fees",
                    "Duties & Brokerage Drawbacks", "Duties & Brokerage",
                    "GST", "EU VAT", "UK VAT"]:
            tot = sum((r[col] or 0) for r in rows)
            totals[col] = round2(tot) if tot else 0
        w.writerow(totals)

    # Bank statement CSV (matches DHLBankStatementImport.csv format)
    out_bank = folder / "dhl_bank_statement.csv"
    today = date.today().strftime("%-m/%-d/%y") if sys.platform != "win32" else date.today().strftime("%#m/%#d/%y")
    tax_total = (category_totals.get("GST", 0)
                 + category_totals.get("EU VAT", 0)
                 + category_totals.get("UK VAT", 0))
    bank_lines = [
        ("Duties & Brokerage",     category_totals.get("Duties & Brokerage", 0)),
        ("GST, EU, UK VAT",        tax_total),
        ("Freight Amounts",        category_totals.get("Freight Amounts", 0)),
        ("Customer Delivery Fees", category_totals.get("Customer Delivery Fees", 0)),
    ]
    with open(out_bank, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["**Date", "**Amount", "Payee", "Description", "Reference", "ChequeNumber"])
        for desc, amt in bank_lines:
            if amt > 0.005:
                w.writerow([today, f"{-round2(amt):.2f}", "DHL", desc, "", ""])

    # Review CSV (only if anything to flag)
    out_review = None
    if reviews:
        out_review = folder / "dhl_review.csv"
        with open(out_review, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["file", "issue"])
            w.writeheader()
            for r in reviews:
                w.writerow(r)

    # Summary to stdout
    grand = round2(sum(r["Amount"] for r in rows))
    print(f"Processed {len(rows)} invoice(s) from {folder.name}")
    print(f"  → {out_reconcile.name}")
    print(f"  → {out_bank.name}")
    if out_review:
        print(f"  ⚠️  {len(reviews)} item(s) flagged → {out_review.name}")
    print()
    print(f"Grand total: ${grand:,.2f}")
    print("Category totals:")
    for cat in ["Freight Amounts", "Customer Delivery Fees",
                "Duties & Brokerage Drawbacks", "Duties & Brokerage",
                "GST", "EU VAT", "UK VAT"]:
        v = category_totals.get(cat, 0)
        if v > 0:
            print(f"  {cat:32} ${v:>10,.2f}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(2)
    main(sys.argv[1])
