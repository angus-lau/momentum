"""
Convert bank activity files to Xero bank statement import format.

Target columns: **Date, **Amount, Payee, Description, Reference, ChequeNumber

Usage:
  python convert_activity.py --all                    # Convert all credit cards (previous month)
  python convert_activity.py --all "26 01"            # Convert all credit cards (specific month)
  python convert_activity.py boa                      # Convert one card (previous month)
  python convert_activity.py boa "26 01"              # Convert one card (specific month)
  python convert_activity.py boa input.csv output.csv # Manual mode
"""

import csv
import glob
import json
import os
import re
import sys
from datetime import date, datetime, timedelta

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

with open(os.path.join(SCRIPT_DIR, "bank_configs.json"), "r") as f:
    BANK_CONFIGS = json.load(f)

with open(os.path.join(SCRIPT_DIR, "upload_configs.json"), "r") as f:
    UPLOAD_CONFIG = json.load(f)

ACCOUNTS = UPLOAD_CONFIG["accounts"]


def get_month_str(month=None):
    """Return 'YY MM' string. Defaults to previous month."""
    if month:
        return month
    first_of_this_month = date.today().replace(day=1)
    prev_month = first_of_this_month - timedelta(days=1)
    return prev_month.strftime("%y %m")


def get_month_folder(bank, month=None):
    """Build the month folder path for a bank account."""
    acct = ACCOUNTS.get(bank)
    if not acct or "folder" not in acct:
        return None
    month_str = get_month_str(month)
    year = 2000 + int(month_str[:2])
    return os.path.join(
        UPLOAD_CONFIG["base_path"],
        f"YE {year}",
        acct["folder"],
        month_str,
    )


def find_activity_file(folder):
    """Find activity.csv, activity.xls, or activity.xlsx in a folder."""
    for ext in ["csv", "xlsx", "xls"]:
        matches = glob.glob(os.path.join(folder, f"activity.{ext}"))
        if matches:
            return matches[0]
    return None


BALANCES_PATH = os.path.join(SCRIPT_DIR, "balances.json")


def load_balances():
    """Load existing balances.json or return empty dict."""
    if os.path.exists(BALANCES_PATH):
        with open(BALANCES_PATH, "r") as f:
            return json.load(f)
    return {}


def save_balances(balances):
    """Save balances dict to balances.json."""
    with open(BALANCES_PATH, "w") as f:
        json.dump(balances, f, indent=2)


def find_statement_pdf(folder):
    """Find a PDF statement in the folder."""
    pdfs = glob.glob(os.path.join(folder, "*.pdf"))
    if len(pdfs) == 1:
        return pdfs[0]
    # If multiple PDFs, look for one with 'statement' in the name
    for pdf in pdfs:
        if "statement" in os.path.basename(pdf).lower():
            return pdf
    return pdfs[0] if pdfs else None


def extract_balance(pdf_path, bank_config_key):
    """Extract ending balance and date from a PDF statement."""
    import pdfplumber

    cfg = BANK_CONFIGS[bank_config_key]
    balance_pattern = cfg.get("balance_pattern")
    date_pattern = cfg.get("date_pattern")

    if not balance_pattern:
        return None, None

    with pdfplumber.open(pdf_path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    balance = None
    closing_date = None

    match = re.search(balance_pattern, text)
    if match:
        balance = float(match.group(1).replace(",", ""))

    if date_pattern:
        match = re.search(date_pattern, text, re.MULTILINE)
        if match:
            raw_date = match.group(1)
            # Normalize to MM/DD/YYYY
            for fmt in ("%m/%d/%Y", "%m/%d/%y", "%b %d, %Y", "%B %d, %Y", "%b%d, %Y"):
                try:
                    closing_date = datetime.strptime(raw_date, fmt).strftime("%m/%d/%Y")
                    break
                except ValueError:
                    continue
            else:
                closing_date = raw_date

    return balance, closing_date


def read_rows_csv(path, cfg):
    """Read rows from a CSV file."""
    rows = []
    with open(path, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for _ in range(cfg.get("skip_rows", 0)):
            next(reader, None)
        if cfg["has_header"]:
            next(reader, None)
        for row in reader:
            rows.append(row)
    return rows


def read_rows_excel(path, cfg):
    """Read rows from an XLS/XLSX file."""
    import openpyxl

    if path.endswith(".xls"):
        # Convert old .xls format — requires xlrd
        import xlrd
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        all_rows = []
        for i in range(ws.nrows):
            all_rows.append([str(ws.cell_value(i, j)) for j in range(ws.ncols)])
    else:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(c) if c is not None else "" for c in row])

    # Apply skip_rows and has_header
    skip = cfg.get("skip_rows", 0)
    all_rows = all_rows[skip:]
    if cfg["has_header"] and all_rows:
        all_rows = all_rows[1:]

    return all_rows


def get_col(row, col_index):
    """Get a column value, supporting single index or list of indices (joined with space)."""
    if isinstance(col_index, list):
        parts = [row[i].strip() for i in col_index if i < len(row) and row[i].strip()]
        return " ".join(parts)
    if col_index < len(row):
        return row[col_index].strip()
    return ""


def write_dhl_reconcile(dhl_amounts, reconcile_path):
    """Write DHL amounts into column A of DHL Reconcile.xlsx, starting at row 3."""
    import openpyxl

    if not os.path.exists(reconcile_path):
        print(f"  DHL Reconcile not found: {reconcile_path}")
        return

    wb = openpyxl.load_workbook(reconcile_path)
    ws = wb.active

    # Find Total row to know where to stop clearing
    total_row = None
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if isinstance(cell_val, str) and cell_val.strip().lower() == "total":
            total_row = row
            break

    # Clear A3 down through old Total row
    clear_to = total_row if total_row else ws.max_row
    for row in range(3, clear_to + 1):
        ws.cell(row=row, column=1).value = None

    # Write DHL amounts in column A starting at row 3
    for i, amount in enumerate(dhl_amounts):
        ws.cell(row=3 + i, column=1).value = amount
        ws.cell(row=3 + i, column=1).number_format = '#,##0.00'

    # Write Total row right after the last amount
    total_r = 3 + len(dhl_amounts)
    last_data_r = total_r - 1
    orange_fill = openpyxl.styles.PatternFill(start_color="FFDEB887", end_color="FFDEB887", fill_type="solid")
    bold_font = openpyxl.styles.Font(bold=True)
    ws.cell(row=total_r, column=1).value = "Total"
    ws.cell(row=total_r, column=1).font = bold_font
    ws.cell(row=total_r, column=1).fill = orange_fill
    # SUM formula for columns B through I
    for col in range(2, 10):  # B=2 through I=9
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.cell(row=total_r, column=col).value = f"=SUM({col_letter}3:{col_letter}{last_data_r})"
        ws.cell(row=total_r, column=col).number_format = '#,##0.00'
        ws.cell(row=total_r, column=col).fill = orange_fill

    wb.save(reconcile_path)
    print(f"  Wrote {len(dhl_amounts)} DHL amounts → {reconcile_path}")


def convert_to_lines(bank_config_key, input_path, dhl_filter=False):
    """Parse a raw activity file into statement line dicts, sign applied per config.

    Returns ``(lines, dhl_amounts)``. Each line is a dict: ``date`` (raw source
    string), ``amount`` (float, sign already resolved via the bank's ``negate`` /
    split-column rule), ``payee``, ``description``, ``reference``, ``cheque``.

    This is the in-memory core shared by ``convert`` (which writes the import CSV)
    and the API upload pipeline (``statement_pipeline.upload_statement``) — so the
    sign convention lives in exactly one place.
    """
    cfg = BANK_CONFIGS[bank_config_key]

    # Read rows based on file type
    if input_path.endswith((".xls", ".xlsx")):
        raw_rows = read_rows_excel(input_path, cfg)
    else:
        raw_rows = read_rows_csv(input_path, cfg)

    lines = []
    dhl_amounts = []

    for row in raw_rows:
        if not row or not row[0].strip():
            continue

        date_val = get_col(row, cfg["date_col"])
        desc = get_col(row, cfg["desc_col"])

        if cfg["amount_type"] == "split":
            debit = get_col(row, cfg["debit_col"])
            credit = get_col(row, cfg["credit_col"])
            if debit:
                amount = -abs(float(debit.replace("$", "").replace(",", "")))
            elif credit:
                amount = abs(float(credit.replace("$", "").replace(",", "")))
            else:
                continue
        else:  # single
            raw = get_col(row, cfg["amount_col"])
            if not raw:
                continue
            amount = float(raw.replace("$", "").replace(",", ""))
            if cfg.get("negate"):
                amount = -amount

        # DHL filtering: separate DHL transactions
        if dhl_filter and "DHL" in desc.upper():
            dhl_amounts.append(amount * -1)
            continue

        lines.append({
            "date": date_val, "amount": amount,
            "payee": desc, "description": desc,
            "reference": "", "cheque": "",
        })

    return lines, dhl_amounts


def convert(bank_config_key, input_path, output_path, dhl_filter=False):
    """Convert an activity file to the Xoro import CSV using the bank config."""
    lines, dhl_amounts = convert_to_lines(bank_config_key, input_path, dhl_filter=dhl_filter)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["**Date", "**Amount", "Payee", "Description", "Reference", "ChequeNumber"])
        for ln in lines:
            writer.writerow([
                ln["date"], f'{ln["amount"]:.2f}',
                ln["payee"], ln["description"], ln["reference"], ln["cheque"],
            ])

    print(f"  Converted {len(lines)} rows ({bank_config_key}) → {output_path}")

    if dhl_filter and dhl_amounts:
        reconcile_path = os.path.join(os.path.dirname(output_path), "DHL Reconcile.xlsx")
        write_dhl_reconcile(dhl_amounts, reconcile_path)

    return len(lines)


def convert_account(bank, month=None):
    """Convert a single account's activity file from its OneDrive folder."""
    acct = ACCOUNTS.get(bank)
    if not acct or "folder" not in acct:
        print(f"  Skipping {bank} — no folder configured")
        return False

    # Determine which bank_config to use (defaults to the account key itself)
    bank_config_key = acct.get("bank_config", bank)
    if bank_config_key not in BANK_CONFIGS:
        print(f"  Skipping {bank} — no bank config for '{bank_config_key}'")
        return False

    folder = get_month_folder(bank, month)
    if not os.path.isdir(folder):
        print(f"  Skipping {bank} — folder not found: {folder}")
        return False

    activity_file = find_activity_file(folder)
    if not activity_file:
        print(f"  Skipping {bank} — no activity file in {folder}")
        return False

    output_path = os.path.join(folder, UPLOAD_CONFIG["filename"])
    dhl_filter = acct.get("dhl_filter", False)
    convert(bank_config_key, activity_file, output_path, dhl_filter=dhl_filter)

    # Extract ending balance from PDF statement
    pdf_file = find_statement_pdf(folder)
    if pdf_file and BANK_CONFIGS[bank_config_key].get("balance_pattern"):
        balance, closing_date = extract_balance(pdf_file, bank_config_key)
        if balance is not None:
            balances = load_balances()
            month_str = get_month_str(month)
            balances.setdefault(month_str, {})[bank] = {
                "balance": balance,
                "date": closing_date,
            }
            save_balances(balances)
            print(f"  Balance: ${balance:,.2f} (closing {closing_date})")

    return True


def convert_all(month=None):
    """Convert all accounts that have folders configured."""
    month_str = get_month_str(month)
    print(f"Converting all accounts for {month_str}...")
    print()

    success = 0
    skipped = 0
    for bank in ACCOUNTS:
        if "folder" not in ACCOUNTS[bank]:
            continue
        print(f"[{bank}]")
        if convert_account(bank, month):
            success += 1
        else:
            skipped += 1
        print()

    print(f"Done: {success} converted, {skipped} skipped")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} --all [month]")
        print(f"       python {sys.argv[0]} <bank> [month]")
        print(f"       python {sys.argv[0]} <bank> <input> [output]")
        print(f"  month format: 'YY MM' (e.g. '26 01')")
        print(f"Banks: {', '.join(BANK_CONFIGS.keys())}")
        sys.exit(1)

    arg1 = sys.argv[1]

    if arg1 == "--all":
        month = sys.argv[2] if len(sys.argv) > 2 else None
        convert_all(month)
    else:
        bank_name = arg1.lower()

        if len(sys.argv) >= 3:
            arg2 = sys.argv[2]
            # Check if it's a month string (e.g. "26 01") or a file path
            if len(arg2) == 5 and arg2[2] == " " and arg2[:2].isdigit() and arg2[3:].isdigit():
                # Month mode
                convert_account(bank_name, arg2)
            else:
                # Manual mode: explicit input/output files
                if bank_name not in BANK_CONFIGS:
                    print(f"Unknown bank config: {bank_name}")
                    print(f"Supported: {', '.join(BANK_CONFIGS.keys())}")
                    sys.exit(1)
                output_file = sys.argv[3] if len(sys.argv) > 3 else "BankStatementImport.csv"
                convert(bank_name, arg2, output_file)
        else:
            # Auto mode: use folder from config, previous month
            convert_account(bank_name)
